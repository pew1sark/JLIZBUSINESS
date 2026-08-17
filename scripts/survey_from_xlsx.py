"""Regenera src/lib/survey.ts desde docs/Levantamiento_Cliente_JLIZ.xlsx.

La planilla Excel es la fuente de verdad de las preguntas: si se agrega o
reformula una pregunta ahí, este script la lleva al formulario web.

Uso:  python3 scripts/survey_from_xlsx.py
"""
import json
from openpyxl import load_workbook

wb = load_workbook("docs/Levantamiento_Cliente_JLIZ.xlsx")
secciones = []
for ws in wb:
    if not ("." in ws.title[:3] and ws.title[0] in "ABCDEFG"):
        continue
    sec = {"key": ws.title[0], "title": ws["A1"].value, "short": ws.title,
           "intro": ws["A2"].value, "blocks": []}
    for r in range(5, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        prio = ws.cell(row=r, column=6).value
        if a and not prio:
            sec["blocks"].append({"title": a, "questions": []})
        elif a and prio:
            sec["blocks"][-1]["questions"].append({
                "id": a,
                "q": ws.cell(row=r, column=2).value,
                "why": ws.cell(row=r, column=3).value,
                "example": ws.cell(row=r, column=4).value,
                "priority": prio,
            })
    secciones.append(sec)

header = open("src/lib/survey.ts").read().split("export const SURVEY")[0]
tail = """

export const TOTAL_QUESTIONS = SURVEY.reduce(
  (n, s) => n + s.blocks.reduce((m, b) => m + b.questions.length, 0),
  0,
)

export const BLOCKING_IDS = new Set(
  SURVEY.flatMap((s) => s.blocks.flatMap((b) => b.questions))
    .filter((q) => q.priority.toLowerCase().includes('bloqueante'))
    .map((q) => q.id),
)
"""
with open("src/lib/survey.ts", "w") as f:
    f.write(header + "export const SURVEY: SurveySection[] = "
            + json.dumps(secciones, ensure_ascii=False, indent=2) + tail)

print("preguntas:", sum(len(b["questions"]) for s in secciones for b in s["blocks"]))
